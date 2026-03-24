"""
blueprints/teacher/routes.py - Teacher Blueprint COMPLETE & FIXED
Handles teacher-specific routes: dashboard, classes, grading, roster management.
This version fixes the routing errors while preserving all existing functionality.
"""

import json
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_required, current_user
from models import Teacher, Class, Enrollment, Student, Subject, Grade, Test
from extensions import db, bcrypt
from config import Config
from sqlalchemy import func
from datetime import datetime, timedelta
from config import Config

# Initialize the blueprint for teacher-related routes
teacher_bp = Blueprint('teacher', __name__)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS — Add these to routes.py (outside any route, at module level)
# ─────────────────────────────────────────────────────────────────────────────
 
def recalculate_term_grade(student_id, class_id, term_tag, teacher_id, commit=True):
    """
    Recalculate and persist the term-level Philippine grade for one student/term.
 
    Uses a synthetic Grade record linked to the FIRST test of the term
    as the "term grade holder". This keeps the data model simple —
    one Grade row per (student, term) rather than a new table.
 
    Returns a dict:
        {
          'grade':      float | None,   # Philippine grade, None = INC
          'display':    str,            # "2.25" or "INC"
          'percentage': float | None,
          'missing':    list[str]       # component names with no scores
        }
    """
    from models import Grade, Test, Class
 
    cls = Class.query.get(class_id)
    if not cls:
        return {'grade': None, 'display': 'INC', 'percentage': None, 'missing': []}
 
    formula = cls.get_grading_formula()
    if not formula or 'components' not in formula:
        return {'grade': None, 'display': 'INC', 'percentage': None, 'missing': ['No formula']}
 
    formula_components = [c['name'] for c in formula['components']]
 
    # ── Check completeness ────────────────────────────────────────────────
    missing = []
    comp_averages = {}
 
    for comp in formula['components']:
        comp_name = comp['name']
        tests_in_group = Test.query.filter_by(
            class_id=class_id,
            term_tag=term_tag,
            component_tag=comp_name
        ).all()
 
        if not tests_in_group:
            missing.append(comp_name)
            continue
 
        percentages = []
        for t in tests_in_group:
            g = Grade.query.filter_by(
                test_id=t.id, student_id=student_id
            ).first()
            if g and g.raw_score is not None and g.max_score and g.max_score > 0:
                percentages.append((g.raw_score / g.max_score) * 100)
 
        if not percentages:
            missing.append(comp_name)
        else:
            comp_averages[comp_name] = sum(percentages) / len(percentages)
 
    if missing:
        # Check if ANY scores have been entered at all for this term
        any_scores = False
        for comp in formula['components']:
            comp_name = comp['name']
            tests_in_group = Test.query.filter_by(
                class_id=class_id,
                term_tag=term_tag,
                component_tag=comp_name
            ).all()
            for t in tests_in_group:
                g = Grade.query.filter_by(
                    test_id=t.id, student_id=student_id
                ).first()
                if g and g.raw_score is not None:
                    any_scores = True
                    break
            if any_scores:
                break

        return {
            'grade':      None,
            'display':    'INC' if any_scores else '—',
            'percentage': None,
            'missing':    missing
        }
 
    # ── Calculate weighted total ──────────────────────────────────────────
    total_weighted = 0.0
    for comp in formula['components']:
        comp_name = comp['name']
        weight    = comp['weight']
        total_weighted += comp_averages[comp_name] * (weight / 100)
 
    percentage = round(total_weighted, 2)
    ph_grade   = cls.convert_to_ph_grade(percentage)
 
    # ── Persist: update enrollment final_grade ────────────────────────────
    # (The enrollment average is handled separately by _update_enrollment_average)
    if commit:
        db.session.commit()
 
    return {
        'grade':      ph_grade,
        'display':    f"{ph_grade:.2f}",
        'percentage': percentage,
        'missing':    []
    }
 
 
def _compute_term_grade_readonly(student_id, class_obj, term_tag):
    """
    Same as recalculate_term_grade but never writes to DB.
    Used when rendering the grading page.
    """
    return recalculate_term_grade(
        student_id=student_id,
        class_id=class_obj.id,
        term_tag=term_tag,
        teacher_id=None,
        commit=False
    )
 
 
def _update_enrollment_average(student_id, class_id):
    """
    Recompute enrollment.final_grade as the average of all
    completed term grades for this student in this class.
    Only terms with a real PH grade (not INC) count.
    Commits to DB.
    """
    from models import Enrollment, Test
 
    cls = Class.query.get(class_id)
    if not cls:
        return
 
    term_tags = (
        db.session.query(Test.term_tag)
        .filter(Test.class_id == class_id, Test.term_tag.isnot(None))
        .distinct()
        .all()
    )
    term_tags = [row[0] for row in term_tags]
 
    completed_grades = []
    for term in term_tags:
        result = recalculate_term_grade(
            student_id=student_id,
            class_id=class_id,
            term_tag=term,
            teacher_id=None,
            commit=False
        )
        if result.get('grade') is not None:
            completed_grades.append(result['grade'])
 
    enrollment = Enrollment.query.filter_by(
        student_id=student_id, class_id=class_id
    ).first()
 
    if enrollment:
        if completed_grades:
            enrollment.final_grade = round(
                sum(completed_grades) / len(completed_grades), 2
            )
        else:
            enrollment.final_grade = None
        db.session.commit()
 

# Decorator to check if current user is a teacher
def teacher_required(f):
    """
    Decorator to ensure only teachers can access the route
    """
    from functools import wraps
    
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_user.role != 'teacher':
            flash('Access denied. Teachers only.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


@teacher_bp.route('/dashboard')
@teacher_required
def dashboard():
    """
    Renders the Teacher Dashboard page with year and semester selection support.
    """
    teacher = current_user.teacher_profile
    
    # Get current values from config
    current_school_year = Config.get_current_school_year()
    current_semester = Config.get_current_semester()
    
    # ✅ Get filters from query parameters
    selected_year = request.args.get('year')
    selected_semester = request.args.get('semester')
    
    # ✅ FIX: Default to current semester if no filter or "all" selected
    if not selected_year or selected_year == 'all':
        selected_year = current_school_year
    
    if not selected_semester or selected_semester == 'all':
        selected_semester = current_semester
    
    # Get ALL classes for this teacher (to build year list)
    all_classes = Class.query.filter_by(teacher_id=teacher.id).all()
    
    # Build available years list
    from datetime import datetime
    actual_years = set([cls.school_year for cls in all_classes])
    current_year_num = datetime.now().year
    buffer_years = {f"{y}-{y+1}" for y in range(current_year_num-5, current_year_num+3)}
    available_years = sorted(list(actual_years | buffer_years), reverse=True)
    
    # ✅ Build query with filters - ALWAYS filter by selected year/semester
    query = Class.query.filter_by(
        teacher_id=teacher.id,
        school_year=selected_year,
        semester=selected_semester
    )
    
    # Execute query
    classes = query.all()
    
    # ✅ Calculate total students for FILTERED classes
    total_students = db.session.query(func.count(func.distinct(Enrollment.student_id)))\
        .join(Class).filter(
            Class.teacher_id == teacher.id,
            Class.school_year == selected_year,
            Class.semester == selected_semester,
            Enrollment.status == 'enrolled'
        ).scalar() or 0
    
    # Calculate weekly schedule for FILTERED classes
    weekly_schedule = {
        'Monday': 0, 'Tuesday': 0, 'Wednesday': 0,
        'Thursday': 0, 'Friday': 0, 'Saturday': 0
    }
    
    for cls in classes:
        if cls.schedule:
            schedule_upper = cls.schedule.upper()
            if 'M' in schedule_upper and 'Monday' not in schedule_upper:
                weekly_schedule['Monday'] += 1
            if 'T' in schedule_upper and 'TH' not in schedule_upper:
                weekly_schedule['Tuesday'] += 1
            if 'W' in schedule_upper:
                weekly_schedule['Wednesday'] += 1
            if 'TH' in schedule_upper or 'R' in schedule_upper:
                weekly_schedule['Thursday'] += 1
            if 'F' in schedule_upper:
                weekly_schedule['Friday'] += 1
            if 'S' in schedule_upper or 'SAT' in schedule_upper:
                weekly_schedule['Saturday'] += 1
    
    # Get classes with stats for FILTERED classes
    classes_with_stats = []
    for cls in classes:
        student_count = Enrollment.query.filter_by(
            class_id=cls.id,
            status='enrolled'
        ).count()
        
        tests = Test.query.filter_by(class_id=cls.id).all()
        ungraded_count = 0
        
        for test in tests:
            enrolled_students = Enrollment.query.filter_by(
                class_id=cls.id,
                status='enrolled'
            ).count()
            
            graded_students = Grade.query.filter_by(
                test_id=test.id
            ).filter(Grade.final_grade.isnot(None)).count()
            
            ungraded_count += (enrolled_students - graded_students)
        
        # Create display dict
        class_display = {
            'id': cls.id,
            'name': cls.effective_subject_name,
            'code': cls.effective_subject_code,
            'section': cls.section
        }
        
        classes_with_stats.append({
            'class': class_display,
            'student_count': student_count,
            'ungraded_count': ungraded_count
        })
    
    return render_template(
        'teacher/dashboard.html',
        teacher=teacher,
        classes_with_stats=classes_with_stats,
        total_students=total_students,
        total_classes=len(classes),
        weekly_schedule=weekly_schedule,
        current_school_year=current_school_year,
        current_semester=current_semester,
        selected_year=selected_year,          # ✅ Pass actual selected value
        selected_semester=selected_semester,  # ✅ Pass actual selected value
        available_years=available_years
    )

@teacher_bp.route('/classes')
@teacher_required
def classes():
    teacher = current_user.teacher_profile
    
    current_school_year = Config.get_current_school_year()
    current_semester = Config.get_current_semester()
    
    selected_year = request.args.get('year')      # None if not in URL
    selected_semester = request.args.get('semester')  # None if not in URL

    # Only default if param was NOT sent at all (first load)
    if selected_year is None:
        selected_year = current_school_year

    if selected_semester is None:
        selected_semester = current_semester

    # Build query - only filter if not 'all'
    query = Class.query.filter_by(teacher_id=teacher.id)
    if selected_year != 'all':
        query = query.filter_by(school_year=selected_year)
    if selected_semester != 'all':
        query = query.filter_by(semester=selected_semester)
    
    # Execute query
    teacher_classes = query.order_by(
        Class.school_year.desc(),
        Class.semester
    ).all()
    
    # Build available years list for dropdown
    all_years = sorted(
        list(set([cls.school_year for cls in Class.query.filter_by(teacher_id=teacher.id).all()])),
        reverse=True
    )
    available_years = all_years  # Show all years with classes
    
    # Get class details with enrollment info and grading progress
    classes_data = []
    for cls in teacher_classes:
        enrollment_count = Enrollment.query.filter_by(
            class_id=cls.id,
            status='enrolled'
        ).count()
        
        # Calculate grading completion percentage
        tests = Test.query.filter_by(class_id=cls.id).all()
        total_possible_grades = enrollment_count * len(tests) if tests else 0
        
        if total_possible_grades > 0:
            graded_count = 0
            for test in tests:
                graded_count += Grade.query.filter_by(
                    test_id=test.id
                ).filter(Grade.final_grade.isnot(None)).count()
            
            grading_percentage = (graded_count / total_possible_grades * 100)
        else:
            grading_percentage = 0
        
        # Create display dict
        class_display = {
            'id': cls.id,
            'name': cls.effective_subject_name,
            'code': cls.effective_subject_code,
            'units': cls.effective_units,
            'section': cls.section,
            'schedule': cls.schedule,
            'room': cls.room,
            'semester': cls.semester,
            'school_year': cls.school_year,
            'max_students': cls.max_students,
            'has_formula': cls.has_grading_formula(),
            'can_edit_formula': cls.can_edit_formula()
        }
        
        classes_data.append({
            'class': class_display,
            'enrollment_count': enrollment_count,
            'capacity_percentage': (enrollment_count / cls.max_students * 100) if cls.max_students else 0,
            'grading_percentage': grading_percentage
        })
    
    # Get all students for the roster builder
    all_students = Student.query.order_by(Student.last_name, Student.first_name).all()
    students_json = [{
        'id': s.id,
        'first_name': s.first_name,
        'last_name': s.last_name,
        'student_number': s.student_number,
        'program': s.program
    } for s in all_students]
    
    return render_template(
        'teacher/classes.html',
        teacher=teacher,
        classes_data=classes_data,
        all_students=students_json,
        current_school_year=current_school_year,
        current_semester=current_semester,
        available_years=available_years,
        selected_year=selected_year or 'all',  # Pass selected filters back
        selected_semester=selected_semester or 'all'
    )


@teacher_bp.route('/classes/create', methods=['POST'])
@teacher_required
def create_class():
    """
    Create a new class with manual subject input and grading formula.
    """
    teacher = current_user.teacher_profile
    
    try:
        # === GET MANUAL SUBJECT INPUT ===
        subject_name = request.form.get('subject_name', '').strip()
        subject_code = request.form.get('subject_code', '').strip()
        units = request.form.get('units', type=int)
        
        # === GET CLASS DETAILS ===
        section = request.form.get('section', '').strip()
        room = request.form.get('room', '').strip()
        schedule = request.form.get('schedule', '').strip()
        school_year = request.form.get('school_year', '').strip()
        semester = request.form.get('semester', '').strip()
        max_students = request.form.get('max_students', type=int, default=40)
        
        # === GET GRADING FORMULA ===
        grading_formula_json = request.form.get('grading_formula', '').strip()
        passing_grade = request.form.get('passing_grade', type=float, default=3.0)
        
        # === GET STUDENT IDS ===
        student_ids_str = request.form.get('student_ids', '')
        
        # === VALIDATE REQUIRED FIELDS ===
        if not subject_name:
            flash('Subject name is required.', 'error')
            return redirect(url_for('teacher.classes'))
        
        if not units or units < 1:
            flash('Units must be at least 1.', 'error')
            return redirect(url_for('teacher.classes'))
        
        if not section or not school_year or not semester:
            flash('Section, school year, and semester are required.', 'error')
            return redirect(url_for('teacher.classes'))
        
        # === VALIDATE GRADING FORMULA ===
        if not grading_formula_json:
            flash('Grading formula is required.', 'error')
            return redirect(url_for('teacher.classes'))
        
        try:
            formula = json.loads(grading_formula_json)
            
            # Validate structure
            if 'components' not in formula or not isinstance(formula['components'], list):
                raise ValueError("Invalid formula structure")
            
            if len(formula['components']) < 1:
                raise ValueError("At least one component is required")
            
            # Validate weights total 100%
            total_weight = sum(c.get('weight', 0) for c in formula['components'])
            if total_weight != 100:
                raise ValueError(f"Component weights must total 100%, got {total_weight}%")
            
            # Validate all components have names
            for comp in formula['components']:
                if not comp.get('name') or comp['name'].strip() == '':
                    raise ValueError("All components must have a name")
                if comp.get('weight', 0) < 0 or comp.get('weight', 0) > 100:
                    raise ValueError("Component weights must be between 0 and 100")
            
        except json.JSONDecodeError:
            flash('Invalid grading formula format.', 'error')
            return redirect(url_for('teacher.classes'))
        except ValueError as e:
            flash(f'Grading formula error: {str(e)}', 'error')
            return redirect(url_for('teacher.classes'))
        
        # === CREATE THE CLASS ===
        new_class = Class(
            teacher_id=teacher.id,
            
            # Manual subject input (NEW)
            subject_name=subject_name,
            subject_code=subject_code or None,
            units=units,
            
            # Class details
            section=section,
            room=room or None,
            schedule=schedule or None,
            school_year=school_year,
            semester=semester,
            max_students=max_students,
            
            # Grading formula (NEW)
            grading_formula=grading_formula_json
        )
        
        db.session.add(new_class)
        db.session.flush()  # Get the class ID
        
        # === ENROLL SELECTED STUDENTS ===
        if student_ids_str:
            student_ids = [int(id) for id in student_ids_str.split(',') if id.strip()]
            
            for student_id in student_ids:
                enrollment = Enrollment(
                    student_id=student_id,
                    class_id=new_class.id,
                    status='enrolled'
                )
                db.session.add(enrollment)
        
        db.session.commit()
        
        # Success message
        display_name = f"{subject_code} - {subject_name}" if subject_code else subject_name
        flash(f'Class "{display_name} (Section {section})" created successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating class: {str(e)}', 'error')
        print(f"Error creating class: {str(e)}")  # For debugging
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('teacher.classes'))


@teacher_bp.route('/classes/delete', methods=['POST'])
@teacher_required
def delete_class():
    """
    Delete a class and all associated enrollments.
    """
    teacher = current_user.teacher_profile
    class_id = request.form.get('class_id', type=int)
    
    # Verify the class belongs to this teacher
    cls = Class.query.filter_by(
        id=class_id,
        teacher_id=teacher.id
    ).first()
    
    if not cls:
        flash('Class not found or access denied.', 'error')
        return redirect(url_for('teacher.classes'))
    
    try:
        class_name = f"{cls.effective_subject_name} - Section {cls.section}"
        
        # Delete the class (cascades to enrollments, tests, and grades)
        db.session.delete(cls)
        db.session.commit()
        
        flash(f'Class "{class_name}" deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting class: {str(e)}', 'error')
    
    return redirect(url_for('teacher.classes'))


@teacher_bp.route('/classes/<int:class_id>')
@teacher_required
def class_detail(class_id):
    """
    Renders the Class Detail page showing roster and class information.
    """
    teacher = current_user.teacher_profile
    
    # Get class and verify it belongs to this teacher
    cls = Class.query.filter_by(
        id=class_id,
        teacher_id=teacher.id
    ).first_or_404()
    
    # Get enrolled students
    enrollments = Enrollment.query.filter_by(
        class_id=class_id,
        status='enrolled'
    ).join(Student).order_by(Student.last_name, Student.first_name).all()
    
    # Get students with their current grades
    students_with_grades = []
    for enrollment in enrollments:
        student = enrollment.student
        
        # Get all grades for this student in this class
        grades = Grade.query.join(Test).filter(
            Test.class_id == class_id,
            Grade.student_id == student.id,
            Grade.final_grade.isnot(None)
        ).all()
        
        # Calculate average grade
        if grades:
            avg_grade = sum(g.final_grade for g in grades) / len(grades)
        else:
            avg_grade = None
        
        students_with_grades.append({
            'student': student,
            'enrollment': enrollment,
            'avg_grade': avg_grade,
            'grade_count': len(grades)
        })
    
    return render_template(
        'teacher/class_detail.html',
        teacher=teacher,
        class_obj=cls,
        students_with_grades=students_with_grades
    )


@teacher_bp.route('/classes/<int:class_id>/formula', methods=['GET'])
@teacher_required
def view_formula(class_id):
    """
    API endpoint to get grading formula for a class
    Used when editing an existing class
    """
    teacher = current_user.teacher_profile
    
    # Verify class belongs to this teacher
    cls = Class.query.filter_by(
        id=class_id,
        teacher_id=teacher.id
    ).first_or_404()
    
    formula = cls.get_grading_formula()
    
    return jsonify({
        'success': True,
        'formula': formula,
        'can_edit': cls.can_edit_formula()
    })


@teacher_bp.route('/classes/<int:class_id>/formula', methods=['POST'])
@teacher_required
def update_formula(class_id):
    """
    API endpoint to update grading formula for a class
    Only allowed if no grades have been entered yet
    """
    teacher = current_user.teacher_profile
    
    # Verify class belongs to this teacher
    cls = Class.query.filter_by(
        id=class_id,
        teacher_id=teacher.id
    ).first_or_404()
    
    # Check if formula can be edited
    if not cls.can_edit_formula():
        return jsonify({
            'success': False,
            'error': 'Cannot edit formula after grades have been entered'
        }), 403
    
    try:
        formula_json = request.json.get('formula')
        
        # Validate and set formula
        cls.set_grading_formula(formula_json)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Formula updated successfully'
        })
        
    except ValueError as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': 'An error occurred while updating the formula'
        }), 500


@teacher_bp.route('/classes/<int:class_id>/enrolled-students', methods=['GET'])
@teacher_required
def get_enrolled_students(class_id):
    """
    API endpoint to get currently enrolled students for a class
    Used in edit mode to show existing roster
    """
    teacher = current_user.teacher_profile
    
    # Verify class belongs to this teacher
    cls = Class.query.filter_by(
        id=class_id,
        teacher_id=teacher.id
    ).first_or_404()
    
    # Get enrolled students
    enrollments = Enrollment.query.filter_by(
        class_id=class_id,
        status='enrolled'
    ).join(Student).order_by(Student.last_name, Student.first_name).all()
    
    students = [{
        'id': e.student.id,
        'first_name': e.student.first_name,
        'last_name': e.student.last_name,
        'student_number': e.student.student_number
    } for e in enrollments]
    
    return jsonify({
        'success': True,
        'students': students
    })


@teacher_bp.route('/classes/<int:class_id>/edit-data', methods=['GET'])
@teacher_required
def get_class_edit_data(class_id):
    """
    API endpoint to get class data for editing
    Returns JSON with class details and what can/cannot be edited
    """
    teacher = current_user.teacher_profile
    
    # Verify class belongs to this teacher
    cls = Class.query.filter_by(
        id=class_id,
        teacher_id=teacher.id
    ).first_or_404()
    
    # Check enrollment count
    enrollment_count = Enrollment.query.filter_by(
        class_id=class_id,
        status='enrolled'
    ).count()
    
    # Check if any grades exist
    has_grades = Grade.query.join(Test).filter(
        Test.class_id == class_id,
        Grade.final_grade.isnot(None)
    ).count() > 0
    
    # Get grading formula
    formula = cls.get_grading_formula()
    
    return jsonify({
        'success': True,
        'data': {
            # Basic info (read-only in edit mode)
            'subject_name': cls.effective_subject_name,
            'subject_code': cls.effective_subject_code,
            'units': cls.effective_units,
            'section': cls.section,
            'school_year': cls.school_year,
            'semester': cls.semester,
            
            # Editable fields
            'schedule': cls.schedule or '',
            'room': cls.room or '',
            'max_students': cls.max_students,
            
            # Formula
            'grading_formula': formula,
            
            # Restrictions
            'enrollment_count': enrollment_count,
            'has_grades': has_grades,
            'can_edit_formula': cls.can_edit_formula()
        }
    })


@teacher_bp.route('/classes/<int:class_id>/update', methods=['POST'])
@teacher_required
def update_class(class_id):
    """
    Update class details (Option A: Safe fields only)
    """
    teacher = current_user.teacher_profile
    
    try:
        # Verify class belongs to this teacher
        cls = Class.query.filter_by(
            id=class_id,
            teacher_id=teacher.id
        ).first_or_404()
        
        # Get form data
        schedule = request.form.get('schedule', '').strip()
        room = request.form.get('room', '').strip()
        max_students = request.form.get('max_students', type=int)
        formula_json = request.form.get('grading_formula', '').strip()
        student_ids_str = request.form.get('student_ids', '')  # ✅ NEW: Get student roster
        
        # ✅ NEW: Parse student IDs
        new_student_ids = set()
        if student_ids_str:
            new_student_ids = set(int(id) for id in student_ids_str.split(',') if id.strip())
        
        # Validate max_students
        if max_students and max_students < 1:
            flash('Maximum students must be at least 1.', 'error')
            return redirect(url_for('teacher.classes'))
        
        # Check enrollment capacity
        enrollment_count = Enrollment.query.filter_by(
            class_id=class_id,
            status='enrolled'
        ).count()
        
        if max_students and max_students < enrollment_count:
            flash(f'Cannot reduce capacity below current enrollment ({enrollment_count} students).', 'error')
            return redirect(url_for('teacher.classes'))
        
        # Update safe fields
        cls.schedule = schedule or None
        cls.room = room or None
        
        if max_students:
            cls.max_students = max_students
        
        # Update formula if allowed
        if formula_json:
            if not cls.can_edit_formula():
                flash('Cannot edit formula after grades have been entered.', 'error')
                return redirect(url_for('teacher.classes'))
            
            try:
                # Validate formula
                formula = json.loads(formula_json)
                
                if 'components' not in formula or not isinstance(formula['components'], list):
                    raise ValueError("Invalid formula structure")
                
                if len(formula['components']) < 1:
                    raise ValueError("At least one component is required")
                
                total_weight = sum(c.get('weight', 0) for c in formula['components'])
                if total_weight != 100:
                    raise ValueError(f"Component weights must total 100%, got {total_weight}%")
                
                for comp in formula['components']:
                    if not comp.get('name') or comp['name'].strip() == '':
                        raise ValueError("All components must have a name")
                    if comp.get('weight', 0) < 0 or comp.get('weight', 0) > 100:
                        raise ValueError("Component weights must be between 0 and 100")
                
                # Set formula
                cls.grading_formula = formula_json
                
            except json.JSONDecodeError:
                flash('Invalid grading formula format.', 'error')
                return redirect(url_for('teacher.classes'))
            except ValueError as e:
                flash(f'Grading formula error: {str(e)}', 'error')
                return redirect(url_for('teacher.classes'))
        
        # ✅ NEW: Update student roster
        # Get currently enrolled students
        current_enrollments = Enrollment.query.filter_by(
            class_id=class_id,
            status='enrolled'
        ).all()
        current_student_ids = set(e.student_id for e in current_enrollments)
        
        # Find students to add (in new list but not in current)
        students_to_add = new_student_ids - current_student_ids
        
        # Find students to remove (in current but not in new list)
        students_to_remove = current_student_ids - new_student_ids
        
        # Add new students
        for student_id in students_to_add:
            # Check if student exists
            student = Student.query.get(student_id)
            if student:
                enrollment = Enrollment(
                    student_id=student_id,
                    class_id=class_id,
                    status='enrolled'
                )
                db.session.add(enrollment)
        
        # Remove students (set status to 'dropped' instead of deleting)
        for student_id in students_to_remove:
            enrollment = Enrollment.query.filter_by(
                class_id=class_id,
                student_id=student_id,
                status='enrolled'
            ).first()
            if enrollment:
                # Check if student has any grades in this class
                has_grades = Grade.query.join(Test).filter(
                    Test.class_id == class_id,
                    Grade.student_id == student_id,
                    Grade.final_grade.isnot(None)
                ).count() > 0
                
                if has_grades:
                    # Don't remove - set to dropped status
                    enrollment.status = 'dropped'
                else:
                    # Safe to delete - no grades exist
                    db.session.delete(enrollment)
        
        db.session.commit()
        
        display_name = f"{cls.effective_subject_code} - {cls.effective_subject_name}"
        flash(f'Class "{display_name}" updated successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating class: {str(e)}', 'error')
        print(f"Error updating class: {str(e)}")
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('teacher.classes'))

@teacher_bp.route('/grading')
@teacher_required
def grading():
    """
    Main grading page. Supports ?year=, ?semester=, ?class_id= params.
 
    grading_data now contains:
        tests           — all Test objects for the class (for column headers)
        students        — list of student data dicts
        term_tags       — ordered list of distinct term tags (for grouping headers)
        formula_components — component names from the formula
 
    Each student dict has:
        student         — Student object
        enrollment      — Enrollment object
        test_grades     — {test.id: Grade or None}   (raw scores per activity)
        term_grades     — {term_tag: {grade, display, percentage}}
    """
    teacher = current_user.teacher_profile
 
    current_school_year = Config.get_current_school_year()
    current_semester    = Config.get_current_semester()
 
    selected_year     = request.args.get('year')
    selected_semester = request.args.get('semester')
 
    if selected_year is None:
        selected_year = current_school_year
    if selected_semester is None:
        selected_semester = current_semester
 
    query = Class.query.filter_by(teacher_id=teacher.id)
    if selected_year != 'all':
        query = query.filter_by(school_year=selected_year)
    if selected_semester != 'all':
        query = query.filter_by(semester=selected_semester)
 
    classes = query.order_by(
        Class.school_year.desc(), Class.semester
    ).all()
 
    all_teacher_classes = Class.query.filter_by(teacher_id=teacher.id).all()
    available_years = sorted(
        list(set([c.school_year for c in all_teacher_classes])), reverse=True
    )
 
    selected_class_id = request.args.get('class_id', type=int)
    if selected_class_id:
        # Only accept this class_id if it's in the filtered list
        selected_class = next((c for c in classes if c.id == selected_class_id), None)
        # If not in the filtered list, fall back to first available
        if not selected_class:
            selected_class = classes[0] if classes else None
    else:
        selected_class = classes[0] if classes else None
 
    grading_data = None
 
    if selected_class:
        enrollments = (
            Enrollment.query
            .filter_by(class_id=selected_class.id, status='enrolled')
            .join(Student)
            .order_by(Student.last_name, Student.first_name)
            .all()
        )
 
        tests = (
            Test.query
            .filter_by(class_id=selected_class.id)
            .order_by(Test.term_tag, Test.component_tag, Test.test_date, Test.created_at)
            .all()
        )
 
        # Collect ordered distinct term tags (preserve entry order)
        seen_terms = {}
        for t in tests:
            if t.term_tag and t.term_tag not in seen_terms:
                seen_terms[t.term_tag] = True
        term_tags = list(seen_terms.keys())
 
        formula = selected_class.get_grading_formula()
        formula_components = [c['name'] for c in formula.get('components', [])] if formula else []
 
        students_data = []
        for enrollment in enrollments:
            student = enrollment.student
 
            # Per-activity raw grades
            test_grades = {}
            for test in tests:
                g = Grade.query.filter_by(
                    test_id=test.id, student_id=student.id
                ).first()
                test_grades[test.id] = g  # full Grade object or None
 
            # Per-term computed grades
            term_grades = {}
            for term in term_tags:
                result = _compute_term_grade_readonly(
                    student_id=student.id,
                    class_obj=selected_class,
                    term_tag=term
                )
                term_grades[term] = result
 
            students_data.append({
                'student':    student,
                'enrollment': enrollment,
                'test_grades': test_grades,
                'term_grades': term_grades
            })
 
        tests_by_term = {}
        for term in term_tags:
            tests_by_term[term] = [t for t in tests if t.term_tag == term]

        grading_data = {
            'students':           students_data,
            'tests':              tests,
            'tests_by_term':      tests_by_term,
            'term_tags':          term_tags,
            'formula_components': formula_components
        }
 
    return render_template(
        'teacher/grading.html',
        teacher=teacher,
        classes=classes,
        selected_class=selected_class,
        grading_data=grading_data,
        current_school_year=current_school_year,
        current_semester=current_semester,
        available_years=available_years,
        selected_year=selected_year or 'all',
        selected_semester=selected_semester or 'all'
    )


@teacher_bp.route('/grading/update', methods=['POST'])
@teacher_required
def update_grade():
    """
    Update a single cell in the spreadsheet.
 
    For tagged (Method 2) tests:
        Expects raw_score and max_score in the POST body.
        Saves them on the Grade record, then triggers recalculate_term_grade()
        which recomputes the term grade for this student.
        The cell value returned is either a Philippine grade or "INC".
 
    For untagged (Method 1 / legacy) tests:
        Accepts a direct grade value as before.
    """
    teacher = current_user.teacher_profile
 
    student_id = request.form.get('student_id', type=int)
    test_id    = request.form.get('test_id', type=int)
 
    test = Test.query.join(Class).filter(
        Test.id == test_id,
        Class.teacher_id == teacher.id
    ).first_or_404()
 
    grade = Grade.query.filter_by(
        test_id=test_id, student_id=student_id
    ).first()
 
    if not grade:
        grade = Grade(
            test_id=test_id, student_id=student_id, graded_by=teacher.id
        )
        db.session.add(grade)
 
    grade.graded_by = teacher.id
    grade.graded_at = datetime.utcnow()
 
    if test.is_tagged:
        # ── Method 2: store raw score, recalculate term ───────────────────
        raw_score = request.form.get('raw_score', type=float)
        max_score = request.form.get('max_score', type=float)
 
        if raw_score is None or max_score is None:
            return jsonify({'success': False, 'error': 'raw_score and max_score required for tagged tests'}), 400
        if max_score <= 0:
            return jsonify({'success': False, 'error': 'max_score must be greater than 0'}), 400
        if raw_score < 0 or raw_score > max_score:
            return jsonify({'success': False, 'error': f'raw_score must be between 0 and {max_score}'}), 400
 
        grade.raw_score = raw_score
        grade.max_score = max_score
        # Individual test cell shows percentage for reference
        grade.calculated_percentage = round((raw_score / max_score) * 100, 2)
        # final_grade on this individual test row = percentage (not PH grade)
        # The PH grade lives on the synthetic term Grade record (see below)
        grade.final_grade = grade.calculated_percentage
 
        db.session.commit()
 
        # Recalculate the term grade for this student
        term_result = recalculate_term_grade(
            student_id=student_id,
            class_id=test.class_id,
            term_tag=test.term_tag,
            teacher_id=teacher.id
        )
 
        # Update enrollment average
        _update_enrollment_average(student_id, test.class_id)
 
        return jsonify({
            'success':            True,
            'grade':              grade.final_grade,          # individual cell value
            'term_tag':           test.term_tag,
            'term_grade':         term_result.get('grade'),   # None = INC
            'term_grade_display': term_result.get('display'), # "2.25" or "INC"
            'term_percentage':    term_result.get('percentage'),
            'missing_components': term_result.get('missing', [])
        })
 
    else:
        # ── Method 1 / legacy: direct grade entry ─────────────────────────
        grade_value = request.form.get('grade', type=float)
        if grade_value is None:
            return jsonify({'success': False, 'error': 'grade value required'}), 400
 
        grade.final_grade      = grade_value
        grade.calculated_grade = grade_value
        db.session.commit()
 
        _update_enrollment_average(student_id, test.class_id)
 
        return jsonify({'success': True, 'grade': grade_value})


@teacher_bp.route('/grading/create-test', methods=['POST'])
@teacher_required
def create_test():
    """
    Create a new test/activity.
 
    Method 2: Accepts term_tag and component_tag so each activity is
    tagged to a grading period and formula component.
 
    POST fields:
        class_id       — required
        title          — required
        term_tag       — required for Method 2 (e.g. "Prelims")
        component_tag  — required for Method 2 (must match a formula component name)
        max_score      — optional, stored on the test for reference
    """
    teacher = current_user.teacher_profile
 
    try:
        class_id      = request.form.get('class_id', type=int)
        title         = request.form.get('title', '').strip()
        term_tag      = request.form.get('term_tag', '').strip() or None
        component_tag = request.form.get('component_tag', '').strip() or None
 
        if not class_id or not title:
            flash('Class and title are required.', 'error')
            return redirect(url_for('teacher.grading'))
 
        cls = Class.query.filter_by(
            id=class_id, teacher_id=teacher.id
        ).first()
        if not cls:
            flash('Class not found or access denied.', 'error')
            return redirect(url_for('teacher.grading'))
 
        # If only one tag is provided, treat as untagged (both required)
        if bool(term_tag) != bool(component_tag):
            flash('Both Term and Component must be provided together.', 'error')
            return redirect(url_for('teacher.grading', class_id=class_id))
 
        # Validate component_tag exists in the formula
        if component_tag:
            formula = cls.get_grading_formula()
            formula_components = [c['name'] for c in formula.get('components', [])]
            if component_tag not in formula_components:
                flash(
                    f'Component "{component_tag}" is not in this class\'s grading formula. '
                    f'Available: {", ".join(formula_components)}',
                    'error'
                )
                return redirect(url_for('teacher.grading', class_id=class_id))
 
        new_test = Test(
            class_id      = class_id,
            title         = title,
            term_tag      = term_tag,
            component_tag = component_tag,
            test_date     = datetime.utcnow().date()
        )
 
        db.session.add(new_test)
        db.session.commit()
 
        tag_info = f" [{term_tag} / {component_tag}]" if term_tag else ""
        flash(f'Test "{title}"{tag_info} created successfully!', 'success')
        return redirect(url_for('teacher.grading', class_id=class_id))
 
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating test: {str(e)}', 'error')
        return redirect(url_for('teacher.grading'))


@teacher_bp.route('/grading/delete-test', methods=['POST'])
@teacher_required
def delete_test():
    """
    Delete a test and all associated grades.
    """
    teacher = current_user.teacher_profile
    
    try:
        test_id = request.form.get('test_id', type=int)
        
        # Verify test belongs to a class taught by this teacher
        test = Test.query.join(Class).filter(
            Test.id == test_id,
            Class.teacher_id == teacher.id
        ).first()
        
        if not test:
            flash('Test not found or access denied.', 'error')
            return redirect(url_for('teacher.grading'))
        
        class_id = test.class_id
        test_title = test.title
        
        # Delete test (cascades to grades)
        db.session.delete(test)
        db.session.commit()
        
        flash(f'Test "{test_title}" deleted successfully.', 'success')
        return redirect(url_for('teacher.grading', class_id=class_id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting test: {str(e)}', 'error')
        return redirect(url_for('teacher.grading'))


# FIX FOR ERROR: "'dict object' has no attribute 'overall_average'"
@teacher_bp.route('/analytics')
@teacher_required
def analytics():
    """
    Renders the Analytics page with comprehensive statistics.
    FIX: Added 'overall_average' key to analytics_data dictionary
    """
    teacher = current_user.teacher_profile
    
    current_school_year = Config.get_current_school_year()
    current_semester = Config.get_current_semester()
    
    # Get all classes for this semester
    classes = Class.query.filter_by(
        teacher_id=teacher.id,
        school_year=current_school_year,
        semester=current_semester
    ).all()
    
    # Calculate total students
    total_students = db.session.query(func.count(func.distinct(Enrollment.student_id)))\
        .join(Class).filter(
            Class.teacher_id == teacher.id,
            Class.school_year == current_school_year,
            Class.semester == current_semester,
            Enrollment.status == 'enrolled'
        ).scalar() or 0
    
    # Calculate new students this semester
    new_students = max(0, total_students - 120)  # Placeholder logic
    
    # Get all grades for this semester
    all_grades = Grade.query.join(Test).join(Class).filter(
        Class.teacher_id == teacher.id,
        Class.school_year == current_school_year,
        Class.semester == current_semester,
        Grade.final_grade.isnot(None)
    ).all()
    
    # Calculate average grade
    if all_grades:
        avg_class_grade = sum(g.final_grade for g in all_grades) / len(all_grades)
    else:
        avg_class_grade = 0.0
    
    # Calculate passing rate
    graded_students = len(all_grades)
    passed_students = sum(1 for g in all_grades if g.final_grade <= 3.0)
    passing_rate = round((passed_students / graded_students * 100) if graded_students > 0 else 0)
    
    # Grade distribution
    grade_distribution = {
        '1.0-1.5': 0,
        '1.75-2.0': 0,
        '2.25-2.5': 0,
        '2.75-3.0': 0,
        '4.0-5.0': 0
    }
    
    for grade in all_grades:
        g = grade.final_grade
        if 1.0 <= g <= 1.5:
            grade_distribution['1.0-1.5'] += 1
        elif 1.75 <= g <= 2.0:
            grade_distribution['1.75-2.0'] += 1
        elif 2.25 <= g <= 2.5:
            grade_distribution['2.25-2.5'] += 1
        elif 2.75 <= g <= 3.0:
            grade_distribution['2.75-3.0'] += 1
        else:
            grade_distribution['4.0-5.0'] += 1
    
    # Class performance breakdown
    class_performance = []
    for cls in classes:
        enrollments = Enrollment.query.filter_by(
            class_id=cls.id,
            status='enrolled'
        ).all()
        
        student_count = len(enrollments)
        
        # Get average grade for this class
        class_grades = Grade.query.join(Test).filter(
            Test.class_id == cls.id,
            Grade.final_grade.isnot(None)
        ).all()
        
        if class_grades:
            avg_grade = sum(g.final_grade for g in class_grades) / len(class_grades)
        else:
            avg_grade = None
        
        class_performance.append({
            'class': cls,
            'student_count': student_count,
            'avg_grade': avg_grade
        })
    
    # Total tests
    total_tests = Test.query.join(Class).filter(
        Class.teacher_id == teacher.id,
        Class.school_year == current_school_year,
        Class.semester == current_semester
    ).count()
    
    # Grading completion
    total_gradable = 0
    graded_count = 0
    
    for cls in classes:
        enrollment_count = Enrollment.query.filter_by(
            class_id=cls.id,
            status='enrolled'
        ).count()
        
        test_count = Test.query.filter_by(class_id=cls.id).count()
        total_gradable += enrollment_count * test_count
        
        graded_count += Grade.query.join(Test).filter(
            Test.class_id == cls.id,
            Grade.final_grade.isnot(None)
        ).count()
    
    completion_rate = round((graded_count / total_gradable * 100) if total_gradable > 0 else 0)
    ungraded_count = total_gradable - graded_count
    
    # Total units
    total_units = sum(cls.units for cls in classes)
    
    # FIX: Added 'overall_average' key - this was causing the error
    analytics_data = {
        'total_students': total_students,
        'new_students': new_students,
        'avg_class_grade': avg_class_grade,
        'overall_average': avg_class_grade,  # ← FIX: This key was missing
        'grade_trend': 0.15,  # Placeholder
        'passing_rate': passing_rate,
        'passed_students': passed_students,
        'graded_students': graded_students,
        'completion_rate': completion_rate,
        'graded_count': graded_count,
        'total_gradable': total_gradable,
        'grade_distribution': grade_distribution,
        'class_performance': class_performance,
        'total_tests': total_tests,
        'avg_tests_per_class': round(total_tests / len(classes), 1) if classes else 0,
        'ungraded_count': ungraded_count,
        'total_classes': len(classes),
        'total_units': total_units
    }
    
    return render_template(
        'teacher/analytics.html',
        teacher=teacher,
        classes=classes,
        analytics=analytics_data,  # Now includes 'overall_average'
        current_school_year=current_school_year,
        current_semester=current_semester
    )


@teacher_bp.route('/profile')
@teacher_required
def profile():
    """
    Renders the Teacher Profile page.
    Shows teacher information and allows profile updates.
    """
    teacher = current_user.teacher_profile
    
    # Get teaching stats
    total_classes = Class.query.filter_by(teacher_id=teacher.id).count()

    current_students = db.session.query(func.count(func.distinct(Enrollment.student_id)))\
        .join(Class).filter(
            Class.teacher_id == teacher.id,
            Class.school_year == Config.get_current_school_year(),
            Class.semester == Config.get_current_semester(),
            Enrollment.status == 'enrolled'
        ).scalar() or 0
    
    # Total grades given
    total_grades_given = Grade.query.join(Test).join(Class).filter(
        Class.teacher_id == teacher.id,
        Grade.final_grade.isnot(None)
    ).count()
    
    # Calculate years teaching (based on account creation)
    years_teaching = max(1, (datetime.utcnow() - teacher.created_at).days // 365) if teacher.created_at else 1
    
    # Recent activity
    recent_grades = Grade.query.join(Test).join(Class).filter(
        Class.teacher_id == teacher.id,
        Grade.graded_at.isnot(None)
    ).order_by(Grade.graded_at.desc()).limit(10).all()
    
    recent_activities = []
    for grade in recent_grades:
        time_diff = datetime.utcnow() - grade.graded_at
        
        if time_diff < timedelta(minutes=60):
            time_ago = f"{time_diff.seconds // 60} minutes ago"
        elif time_diff < timedelta(hours=24):
            time_ago = f"{time_diff.seconds // 3600} hours ago"
        else:
            time_ago = f"{time_diff.days} days ago"
        
        recent_activities.append({
            'description': f"Graded {grade.student.get_full_name()} in {grade.test.class_.effective_subject_code}",
            'time_ago': time_ago
        })
            
    # If no recent grading activity, show placeholder
    if not recent_activities:
        recent_activities = [
            {'description': 'No recent grading activity', 'time_ago': 'Start grading to see activity here'}
        ]
    
    stats = {
        'total_classes': total_classes,
        'current_students': current_students,
        'total_grades_given': total_grades_given,
        'years_teaching': years_teaching
    }
    
    return render_template(
        'teacher/profile.html',
        teacher=teacher,
        stats=stats,
        recent_activities=recent_activities
    )

@teacher_bp.route('/change-password', methods=['POST'])
@teacher_required
def change_password():
    """Change teacher's password"""
    try:
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validation
        if not all([current_password, new_password, confirm_password]):
            flash('All fields are required.', 'error')
            return redirect(url_for('teacher.profile'))
        
        # Check if new passwords match
        if new_password != confirm_password:
            flash('New passwords do not match.', 'error')
            return redirect(url_for('teacher.profile'))
        
        # Check password length
        if len(new_password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return redirect(url_for('teacher.profile'))
        
        # Verify current password
        if not bcrypt.check_password_hash(current_user.password, current_password):
            flash('Current password is incorrect.', 'error')
            return redirect(url_for('teacher.profile'))
        
        # Update password
        current_user.password = bcrypt.generate_password_hash(new_password).decode('utf-8')
        db.session.commit()
        
        flash('✅ Password updated successfully!', 'success')
        return redirect(url_for('teacher.profile'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error updating password: {str(e)}', 'error')
        return redirect(url_for('teacher.profile'))


"""
NEW ROUTES TO ADD TO blueprints/teacher/routes.py
Add these routes to your existing routes.py file
"""

@teacher_bp.route('/grading/components', methods=['GET'])
@teacher_required
def get_components():
    """
    API endpoint to get component scores for a student/test
    """
    teacher = current_user.teacher_profile
    
    student_id = request.args.get('student_id', type=int)
    test_id = request.args.get('test_id', type=int)
    
    # Verify test belongs to teacher's class
    test = Test.query.join(Class).filter(
        Test.id == test_id,
        Class.teacher_id == teacher.id
    ).first_or_404()
    
    # Get existing grade
    grade = Grade.query.filter_by(
        test_id=test_id,
        student_id=student_id
    ).first()
    
    if grade:
        components = grade.get_component_scores()
    else:
        components = {}
    
    return jsonify({
        'success': True,
        'components': components
    })


@teacher_bp.route('/grading/update-components', methods=['POST'])
@teacher_required
def update_components():
    """
    API endpoint to save component scores (Option B)
    Automatically calculates final grade
    """
    teacher = current_user.teacher_profile
    
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        test_id = data.get('test_id')
        components = data.get('components', {})
        
        # Verify test belongs to teacher's class
        test = Test.query.join(Class).filter(
            Test.id == test_id,
            Class.teacher_id == teacher.id
        ).first_or_404()
        
        # Get or create grade
        grade = Grade.query.filter_by(
            test_id=test_id,
            student_id=student_id
        ).first()
        
        if not grade:
            grade = Grade(
                test_id=test_id,
                student_id=student_id,
                graded_by=teacher.id
            )
            db.session.add(grade)
        
        # Save component scores
        grade.set_component_scores(components)
        
        # Calculate final grade using Option B logic
        grade.calculate_grade(test.class_)
        
        grade.graded_by = teacher.id
        grade.graded_at = datetime.utcnow()
        
        db.session.commit()
        
        # Update enrollment final grade (average of all tests)
        update_enrollment_average(student_id, test.class_id)
        
        return jsonify({
            'success': True,
            'final_grade': grade.final_grade,
            'calculated_percentage': grade.calculated_percentage
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@teacher_bp.route('/grading/student-average', methods=['GET'])
@teacher_required
def get_student_average():
    """
    Returns the student's current average AND all term grades for a class.
    Called after every grade save to refresh the spreadsheet.
 
    Response:
    {
        "success": true,
        "average": 2.25,             // enrollment final_grade (or null)
        "average_display": "2.25",   // string, "INC" if null
        "term_grades": {
            "Prelims":  {"grade": 2.25, "display": "2.25", "percentage": 83.2},
            "Midterms": {"grade": null, "display": "INC",  "percentage": null},
            "Finals":   {"grade": null, "display": "INC",  "percentage": null}
        }
    }
    """
    teacher    = current_user.teacher_profile
    student_id = request.args.get('student_id', type=int)
    class_id   = request.args.get('class_id', type=int)
 
    cls = Class.query.filter_by(
        id=class_id, teacher_id=teacher.id
    ).first_or_404()
 
    # Collect all distinct term_tags used in this class
    term_tags = (
        db.session.query(Test.term_tag)
        .filter(Test.class_id == class_id, Test.term_tag.isnot(None))
        .distinct()
        .all()
    )
    term_tags = [row[0] for row in term_tags]
 
    term_grades = {}
    completed_ph_grades = []
 
    for term in term_tags:
        result = recalculate_term_grade(
            student_id=student_id,
            class_id=class_id,
            term_tag=term,
            teacher_id=teacher.id,
            commit=False   # don't commit during a read-only request
        )
        term_grades[term] = result
        if result.get('grade') is not None:
            completed_ph_grades.append(result['grade'])
 
    # Overall average: average of completed terms only
    if completed_ph_grades:
        avg = round(sum(completed_ph_grades) / len(completed_ph_grades), 2)
        avg_display = f"{avg:.2f}"
    else:
        # Check if any term has started (INC) vs nothing entered at all (—)
        any_inc = any(
            tg.get('display') == 'INC' 
            for tg in term_grades.values()
        )
        avg = None
        avg_display = "INC" if any_inc else "—"
 
    return jsonify({
        'success':         True,
        'average':         avg,
        'average_display': avg_display,
        'term_grades':     term_grades
    })


def calculate_class_average(student_id, class_id):
    """
    Helper function to calculate student's average grade for a class
    """
    # Get all grades for this student in this class
    all_grades = Grade.query.join(Test).filter(
        Test.class_id == class_id,
        Grade.student_id == student_id,
        Grade.final_grade.isnot(None)
    ).all()
    
    if not all_grades:
        return None
    
    # Simple average of all test grades
    total = sum(g.final_grade for g in all_grades)
    return round(total / len(all_grades), 2)


def update_enrollment_average(student_id, class_id):
    """
    Helper function to update enrollment average after grade change
    """
    enrollment = Enrollment.query.filter_by(
        student_id=student_id,
        class_id=class_id
    ).first()
    
    if enrollment:
        average = calculate_class_average(student_id, class_id)
        enrollment.final_grade = average
        db.session.commit()


@teacher_bp.route('/grading/export')
@teacher_required
def export_grades():
    """
    Export grades for a class as Excel (.xlsx) or CSV (.csv).
 
    Query params:
        class_id  — required
        format    — 'xlsx' or 'csv' (default: xlsx)
    """
    import io
    import csv
    from flask import send_file, Response
 
    teacher = current_user.teacher_profile
 
    class_id     = request.args.get('class_id', type=int)
    fmt          = request.args.get('format', 'xlsx').lower()
 
    if not class_id:
        flash('No class selected for export.', 'error')
        return redirect(url_for('teacher.grading'))
 
    cls = Class.query.filter_by(
        id=class_id, teacher_id=teacher.id
    ).first_or_404()
 
    # ── Gather data ──────────────────────────────────────────────────────────
 
    enrollments = (
        Enrollment.query
        .filter_by(class_id=class_id, status='enrolled')
        .join(Student)
        .order_by(Student.last_name, Student.first_name)
        .all()
    )
 
    tests = (
        Test.query
        .filter_by(class_id=class_id)
        .order_by(Test.term_tag, Test.component_tag, Test.test_date, Test.created_at)
        .all()
    )
 
    # Ordered distinct term tags
    seen_terms = {}
    for t in tests:
        if t.term_tag and t.term_tag not in seen_terms:
            seen_terms[t.term_tag] = True
    term_tags = list(seen_terms.keys())
 
    tests_by_term = {term: [t for t in tests if t.term_tag == term] for term in term_tags}
 
    # Class info
    subject_name = cls.effective_subject_name
    subject_code = cls.effective_subject_code
    section      = cls.section or ''
    school_year  = cls.school_year
    semester     = cls.semester
 
    safe_name = f"{subject_code}_{section}_{school_year}_{semester}".replace(' ', '_').replace('/', '-')
 
    # ── Build rows ───────────────────────────────────────────────────────────
 
    # Header row 1: info labels
    # Header row 2: column names
    # Data rows: one per student
 
    col_headers = ['Student Name', 'Student ID']
 
    if term_tags:
        for term in term_tags:
            for t in tests_by_term[term]:
                col_headers.append(f'{term} | {t.title}')
            col_headers.append(f'{term} Grade')
    else:
        for t in tests:
            col_headers.append(t.title)
 
    col_headers.append('Overall Grade')
 
    data_rows = []
    for enrollment in enrollments:
        student = enrollment.student
 
        row = [student.get_full_name(), student.student_number]
 
        if term_tags:
            for term in term_tags:
                for t in tests_by_term[term]:
                    g = Grade.query.filter_by(
                        test_id=t.id, student_id=student.id
                    ).first()
                    if g and g.raw_score is not None:
                        max_s = g.max_score if g.max_score else '?'
                        row.append(f"{g.raw_score}/{max_s}")
                    else:
                        row.append('')
 
                # Term grade
                result = recalculate_term_grade(
                    student_id=student.id,
                    class_id=class_id,
                    term_tag=term,
                    teacher_id=None,
                    commit=False
                )
                row.append(result.get('display', '—'))
        else:
            for t in tests:
                g = Grade.query.filter_by(
                    test_id=t.id, student_id=student.id
                ).first()
                if g and g.final_grade is not None:
                    row.append(f"{g.final_grade:.2f}")
                else:
                    row.append('')
 
        # Overall
        if enrollment.final_grade is not None:
            row.append(f"{enrollment.final_grade:.2f}")
        else:
            row.append('INC' if any(
                recalculate_term_grade(student.id, class_id, term, None, False).get('display') == 'INC'
                for term in term_tags
            ) else '—')
 
        data_rows.append(row)
 
    # ── CSV export ───────────────────────────────────────────────────────────
 
    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
 
        # Class info header block
        writer.writerow(['Subject', subject_name])
        writer.writerow(['Code', subject_code])
        writer.writerow(['Section', section])
        writer.writerow(['School Year', school_year])
        writer.writerow(['Semester', semester])
        writer.writerow([])  # blank separator
 
        writer.writerow(col_headers)
        for row in data_rows:
            writer.writerow(row)
 
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=grades_{safe_name}.csv'
            }
        )
 
    # ── Excel export ─────────────────────────────────────────────────────────
 
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
 
    wb = Workbook()
    ws = wb.active
    ws.title = 'Grades'
 
    # Styles
    header_font      = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    info_font        = Font(name='Arial', bold=True, size=10)
    label_font       = Font(name='Arial', bold=True, size=9)
    data_font        = Font(name='Arial', size=9)
    header_fill      = PatternFill('solid', start_color='1F3864')   # dark navy
    term_fill        = PatternFill('solid', start_color='2E4D7B')   # medium navy
    term_grade_fill  = PatternFill('solid', start_color='3A5A8C')   # lighter navy
    col_header_fill  = PatternFill('solid', start_color='4472C4')   # blue
    overall_fill     = PatternFill('solid', start_color='375623')   # dark green
    alt_row_fill     = PatternFill('solid', start_color='EBF3FF')   # light blue tint
    center           = Alignment(horizontal='center', vertical='center')
    left             = Alignment(horizontal='left',   vertical='center')
    thin             = Side(style='thin', color='BFBFBF')
    border           = Border(left=thin, right=thin, top=thin, bottom=thin)
 
    # ── Row 1–5: Class info block ─────────────────────────────────────────
    info_pairs = [
        ('Subject',     subject_name),
        ('Code',        subject_code),
        ('Section',     section),
        ('School Year', school_year),
        ('Semester',    semester),
    ]
    for i, (label, value) in enumerate(info_pairs, start=1):
        ws.cell(row=i, column=1, value=label).font = info_font
        cell = ws.cell(row=i, column=2, value=value)
        cell.font = Font(name='Arial', size=10)
    
    info_rows = len(info_pairs)
    blank_row  = info_rows + 1
    ws.append([])  # blank separator
 
    # ── Term group header row (only if tagged tests exist) ────────────────
    current_row = blank_row + 1
    col_cursor  = 3  # columns 1=Name, 2=ID, then activities
 
    if term_tags:
        # Row: merged term label spanning activity cols + 1 grade col
        for r in range(1, current_row):
            ws.cell(row=r, column=1).fill = PatternFill('solid', start_color='F2F2F2')
 
        # Name + ID headers (merged across group header row)
        name_cell = ws.cell(row=current_row, column=1, value='Student Name')
        name_cell.font  = header_font
        name_cell.fill  = header_fill
        name_cell.alignment = center
        name_cell.border = border
 
        id_cell = ws.cell(row=current_row, column=2, value='Student ID')
        id_cell.font  = header_font
        id_cell.fill  = header_fill
        id_cell.alignment = center
        id_cell.border = border
 
        for term in term_tags:
            term_tests    = tests_by_term[term]
            span          = len(term_tests) + 1   # activities + 1 grade col
            start_col     = col_cursor
            end_col       = col_cursor + span - 1
 
            # Merge term label
            if span > 1:
                ws.merge_cells(
                    start_row=current_row, start_column=start_col,
                    end_row=current_row,   end_column=end_col
                )
            cell = ws.cell(row=current_row, column=start_col, value=term)
            cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            cell.fill      = term_fill
            cell.alignment = center
            cell.border    = border
 
            col_cursor += span
 
        # Overall column header (group row)
        ov_cell = ws.cell(row=current_row, column=col_cursor, value='Overall')
        ov_cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        ov_cell.fill      = overall_fill
        ov_cell.alignment = center
        ov_cell.border    = border
 
        current_row += 1
 
    # ── Activity-level column headers ────────────────────────────────────
    ws.cell(row=current_row, column=1, value='Student Name').font  = label_font
    ws.cell(row=current_row, column=1).fill      = col_header_fill
    ws.cell(row=current_row, column=1).alignment = left
    ws.cell(row=current_row, column=1).font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    ws.cell(row=current_row, column=1).border    = border
 
    ws.cell(row=current_row, column=2, value='Student ID').fill      = col_header_fill
    ws.cell(row=current_row, column=2).alignment = center
    ws.cell(row=current_row, column=2).font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    ws.cell(row=current_row, column=2).border    = border
 
    col_cursor = 3
    if term_tags:
        for term in term_tags:
            for t in tests_by_term[term]:
                cell = ws.cell(row=current_row, column=col_cursor, value=t.title)
                cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=8)
                cell.fill      = col_header_fill
                cell.alignment = center
                cell.border    = border
                col_cursor += 1
            # Term grade column header
            cell = ws.cell(row=current_row, column=col_cursor, value='Grade')
            cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
            cell.fill      = term_grade_fill
            cell.alignment = center
            cell.border    = border
            col_cursor += 1
    else:
        for t in tests:
            cell = ws.cell(row=current_row, column=col_cursor, value=t.title)
            cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
            cell.fill      = col_header_fill
            cell.alignment = center
            cell.border    = border
            col_cursor += 1
 
    # Overall
    cell = ws.cell(row=current_row, column=col_cursor, value='Overall Grade')
    cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    cell.fill      = overall_fill
    cell.alignment = center
    cell.border    = border
 
    current_row += 1
    total_cols = col_cursor
 
    # ── Data rows ─────────────────────────────────────────────────────────
    for idx, row_values in enumerate(data_rows):
        is_alt = (idx % 2 == 1)
        for col_i, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_i, value=value)
            cell.font   = data_font
            cell.border = border
            if col_i == 1:
                cell.alignment = left
            else:
                cell.alignment = center
            if is_alt:
                cell.fill = alt_row_fill
            # Highlight INC in orange
            if value == 'INC':
                cell.font = Font(name='Arial', size=9, bold=True, color='D29922')
        current_row += 1
 
    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 28   # Student Name
    ws.column_dimensions['B'].width = 16   # Student ID
    for col_i in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 14
 
    # ── Row heights ───────────────────────────────────────────────────────
    for r in range(1, current_row):
        ws.row_dimensions[r].height = 18
 
    # ── Freeze panes: freeze after name + ID columns ──────────────────────
    freeze_row = blank_row + 2 if term_tags else blank_row + 1
    ws.freeze_panes = ws.cell(row=freeze_row + 1, column=3)
 
    # ── Stream to client ──────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
 
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'grades_{safe_name}.xlsx'
    )